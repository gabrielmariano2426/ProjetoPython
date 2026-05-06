# view.py

from modelo import Conta, Historico, Tipos, Status, Bancos, engine
from sqlmodel import Session, select
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def criar_contas(conta: Conta):
    with Session(engine) as session:
        statement = select(Conta).where(Conta.banco == conta.banco)
        results = session.exec(statement).all()
        if results:
            print('Já existe uma conta neste banco!')
            return
        session.add(conta)
        session.commit()
        return conta

def listar_contas():
    with Session(engine) as session:
        statement = select(Conta)
        results = session.exec(statement).all()
        return results

def desativar_conta(id):
    with Session(engine) as session:
        statement = select(Conta).where(Conta.id == id)
        conta = session.exec(statement).first()
        if conta.valor > 0:
            raise ValueError('Essa conta ainda possui saldo, não é possível desativar')
        conta.status = Status.INATIVO.value
        session.commit()

def transferir_saldo(id_conta_saida, id_conta_entrada, valor):
    with Session(engine) as session:
        statement = select(Conta).where(Conta.id == id_conta_saida)
        conta_saida = session.exec(statement).first()
        if conta_saida.valor < valor:
            raise ValueError('Saldo insuficiente')
        statement = select(Conta).where(Conta.id == id_conta_entrada)
        conta_entrada = session.exec(statement).first()
        conta_saida.valor -= valor
        conta_entrada.valor += valor
        session.commit()

def movimentar_dinheiro(historico: Historico):
    with Session(engine) as session:
        statement = select(Conta).where(Conta.id == historico.conta_id)
        conta = session.exec(statement).first()

        # ✅ Verifica se a conta existe
        if not conta:
            raise ValueError("Conta não encontrada!")

        # ✅ Verifica se a conta está ativa
        if conta.status == Status.INATIVO.value:
            raise ValueError("Não é possível movimentar uma conta inativa!")

        if historico.tipo == Tipos.ENTRADA.value:
            conta.valor += historico.valor
        else:
            if conta.valor < historico.valor:
                raise ValueError("Saldo insuficiente")
            conta.valor -= historico.valor

        session.add(historico)
        session.commit()
        return historico

def total_contas():
    with Session(engine) as session:
        statement = select(Conta)
        contas = session.exec(statement).all()
    total = 0
    for conta in contas:
        total += conta.valor
    return float(total)

def buscar_historicos_entre_datas(data_inicio: date, data_fim: date):
    with Session(engine) as session:
        statement = select(Historico).where(
            Historico.data >= data_inicio,
            Historico.data <= data_fim
        )
        resultados = session.exec(statement).all()
        return resultados

def criar_grafico_por_conta():
    with Session(engine) as session:
        statement = select(Conta).where(Conta.status == Status.ATIVO.value)
        contas = session.exec(statement).all()
        bancos = [i.banco for i in contas]
        total = [i.valor for i in contas]
    import matplotlib.pyplot as plt
    plt.bar(bancos, total)
    plt.show()

def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # ── Aba Contas ──────────────────────────────────────────
    ws_contas = wb.active
    ws_contas.title = 'Contas'

    headers_contas = ['ID', 'Banco', 'Status', 'Saldo (R$)', 'Total Entradas (R$)', 'Total Saídas (R$)']
    cor_azul = '4472C4'
    for col, header in enumerate(headers_contas, 1):
        cell = ws_contas.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial')
        cell.fill = PatternFill('solid', start_color=cor_azul)
        cell.alignment = Alignment(horizontal='center')

    with Session(engine) as session:
        contas = session.exec(select(Conta)).all()
        historicos = session.exec(select(Historico)).all()

    # Agrupa históricos por conta
    from collections import defaultdict
    entradas = defaultdict(float)
    saidas = defaultdict(float)
    for h in historicos:
        if h.tipo == Tipos.ENTRADA.value:
            entradas[h.conta_id] += h.valor
        else:
            saidas[h.conta_id] += h.valor

    for row, conta in enumerate(contas, 2):
        ws_contas.cell(row=row, column=1, value=conta.id)
        ws_contas.cell(row=row, column=2, value=conta.banco)
        ws_contas.cell(row=row, column=3, value=conta.status)
        ws_contas.cell(row=row, column=4, value=conta.valor)
        ws_contas.cell(row=row, column=5, value=entradas[conta.id])
        ws_contas.cell(row=row, column=6, value=saidas[conta.id])

    # Linha de totais
    last = len(contas) + 2
    total_cell = ws_contas.cell(row=last, column=3, value='TOTAL')
    total_cell.font = Font(bold=True, name='Arial')
    for col in [4, 5, 6]:
        cell = ws_contas.cell(row=last, column=col, value=f'=SUM({chr(64+col)}2:{chr(64+col)}{last-1})')
        cell.font = Font(bold=True, name='Arial')

    for col_letter, width in zip(['A','B','C','D','E','F'], [8, 14, 12, 16, 22, 20]):
        ws_contas.column_dimensions[col_letter].width = width

    # ── Uma aba por conta com suas movimentações ────────────
    cor_verde = '70AD47'
    cor_laranja = 'ED7D31'

    for conta in contas:
        nome_aba = f'Conta {conta.id} - {conta.banco}'[:31]  # Excel limita 31 chars
        ws = wb.create_sheet(nome_aba)

        # Cabeçalho da conta
        ws.merge_cells('A1:E1')
        titulo = ws['A1']
        titulo.value = f'{conta.banco} | Saldo: R$ {conta.valor:.2f} | Status: {conta.status}'
        titulo.font = Font(bold=True, color='FFFFFF', name='Arial', size=12)
        titulo.fill = PatternFill('solid', start_color=cor_azul)
        titulo.alignment = Alignment(horizontal='center')

        # Headers das movimentações
        headers_mov = ['ID', 'Tipo', 'Valor (R$)', 'Data', 'Saldo Acumulado (R$)']
        for col, header in enumerate(headers_mov, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF', name='Arial')
            cell.fill = PatternFill('solid', start_color=cor_verde)
            cell.alignment = Alignment(horizontal='center')

        # Movimentações desta conta ordenadas por data
        movs = sorted([h for h in historicos if h.conta_id == conta.id], key=lambda x: x.data)

        saldo_acumulado = 0
        for row, h in enumerate(movs, 3):
            if h.tipo == Tipos.ENTRADA.value:
                saldo_acumulado += h.valor
                cor_linha = 'E2EFDA'  # verde claro para entrada
            else:
                saldo_acumulado -= h.valor
                cor_linha = 'FCE4D6'  # laranja claro para saída

            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill('solid', start_color=cor_linha)

            ws.cell(row=row, column=1, value=h.id)
            ws.cell(row=row, column=2, value=h.tipo)
            ws.cell(row=row, column=3, value=h.valor)
            ws.cell(row=row, column=4, value=str(h.data))
            ws.cell(row=row, column=5, value=round(saldo_acumulado, 2))

        # Rodapé com totais da conta
        if movs:
            last_row = len(movs) + 3
            ws.cell(row=last_row, column=1, value='TOTAL').font = Font(bold=True, name='Arial')
            cell_entrada = ws.cell(row=last_row, column=2, value=f'Entradas: R$ {entradas[conta.id]:.2f}')
            cell_entrada.font = Font(bold=True, color='375623', name='Arial')
            cell_saida = ws.cell(row=last_row, column=3, value=f'Saídas: R$ {saidas[conta.id]:.2f}')
            cell_saida.font = Font(bold=True, color='833C00', name='Arial')

        for col_letter, width in zip(['A','B','C','D','E'], [8, 12, 14, 14, 22]):
            ws.column_dimensions[col_letter].width = width

    # ── Aba Histórico Geral ──────────────────────────────────
    ws_hist = wb.create_sheet('Histórico Geral')

    headers_hist = ['ID', 'Conta ID', 'Banco', 'Tipo', 'Valor (R$)', 'Data']
    for col, header in enumerate(headers_hist, 1):
        cell = ws_hist.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial')
        cell.fill = PatternFill('solid', start_color=cor_laranja)
        cell.alignment = Alignment(horizontal='center')

    conta_map = {c.id: c.banco for c in contas}
    historicos_ord = sorted(historicos, key=lambda x: x.data)

    for row, h in enumerate(historicos_ord, 2):
        cor_linha = 'E2EFDA' if h.tipo == Tipos.ENTRADA.value else 'FCE4D6'
        for col in range(1, 7):
            ws_hist.cell(row=row, column=col).fill = PatternFill('solid', start_color=cor_linha)

        ws_hist.cell(row=row, column=1, value=h.id)
        ws_hist.cell(row=row, column=2, value=h.conta_id)
        ws_hist.cell(row=row, column=3, value=conta_map.get(h.conta_id, '?'))
        ws_hist.cell(row=row, column=4, value=h.tipo)
        ws_hist.cell(row=row, column=5, value=h.valor)
        ws_hist.cell(row=row, column=6, value=str(h.data))

    for col_letter, width in zip(['A','B','C','D','E','F'], [8, 10, 14, 12, 14, 14]):
        ws_hist.column_dimensions[col_letter].width = width

    wb.save('relatorio.xlsx')
    print('✅ Arquivo relatorio.xlsx gerado com sucesso!')